"""Attendance Management System — Flask Application"""
import csv
import io
import json as _json
import os
import re
import secrets
import threading as _threading
import urllib.request
import urllib.error
from datetime import datetime as _dt, timedelta
from functools import wraps

from dotenv import load_dotenv
load_dotenv()

from flask import (Flask, Response, flash, jsonify, redirect,
                   render_template, request, session, url_for, send_file)
from flask_wtf.csrf import CSRFProtect
from werkzeug.security import check_password_hash, generate_password_hash

import db
import processor

app = Flask(__name__)

# ── Security configuration ─────────────────────────────────────────
_FORCE_HTTPS = os.environ.get('FORCE_HTTPS', 'false').lower() == 'true'
app.config.update(
    SECRET_KEY              = os.environ.get('SECRET_KEY', secrets.token_hex(32)),
    SESSION_COOKIE_HTTPONLY = True,
    SESSION_COOKIE_SAMESITE = 'Lax',
    SESSION_COOKIE_SECURE   = _FORCE_HTTPS,
    PERMANENT_SESSION_LIFETIME = timedelta(minutes=30),
    WTF_CSRF_TIME_LIMIT     = None,
    MAX_CONTENT_LENGTH      = 10 * 1024 * 1024,   # 10 MB upload limit
)
csrf = CSRFProtect(app)

# Fernet encryption for export files at rest
try:
    from cryptography.fernet import Fernet as _Fernet
    _fk = os.environ.get('FERNET_KEY', '')
    _fernet = _Fernet(_fk.encode()) if _fk else None
except Exception:
    _fernet = None

def _encrypt(data: bytes) -> bytes:
    return _fernet.encrypt(data) if _fernet else data

def _decrypt(data: bytes) -> bytes:
    return _fernet.decrypt(data) if _fernet else data

# Directory for persisted export files
EXPORTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')
os.makedirs(EXPORTS_DIR, exist_ok=True)

db.init_db()


# ------------------------------------------------------------------ #
# EXPORT HELPERS
# ------------------------------------------------------------------ #

def _resolve_export_dates(filters):
    """Return (from_date, to_date) strings from filter dict."""
    import calendar as _cal
    month  = filters.get('month')
    period = filters.get('period')
    if month:
        yr, mo = map(int, month.split('-'))
        last   = _cal.monthrange(yr, mo)[1]
        if period == '1':   return f'{month}-01', f'{month}-15'
        elif period == '2': return f'{month}-16', f'{month}-{last:02d}'
        else:               return f'{month}-01', f'{month}-{last:02d}'
    return filters.get('from_date') or '', filters.get('to_date') or ''


def _export_fname(label, from_d, to_d, ext):
    """Build a human-readable, filesystem-safe filename."""
    def _fmt(d):
        try:
            y, m, day = d.split('-')
            return f'{day}-{m}-{y}'
        except Exception:
            return d
    if from_d and to_d:
        return f'Attendance_Report_{_fmt(from_d)}_den_{_fmt(to_d)}.{ext}'
    return f'Attendance_Report_Toan_Bo_{_dt.now().strftime("%d-%m-%Y_%H%M")}.{ext}'


def _persist_export(raw_bytes, filename, export_type, from_d, to_d):
    """Encrypt and save bytes to exports/ dir, then record in DB."""
    filepath = os.path.join(EXPORTS_DIR, filename)
    with open(filepath, 'wb') as f:
        f.write(_encrypt(raw_bytes))
    db.save_export_record({
        'filename':    filename,
        'filepath':    filepath,
        'export_type': export_type,
        'date_from':   from_d,
        'date_to':     to_d,
        'exported_by': session.get('display_name') or session.get('username', ''),
        'file_size':   os.path.getsize(filepath),
    })
    db.log_audit(session.get('user_id', 0), session.get('username', ''), 'EXPORT',
                 f'{export_type} | {filename}', request.remote_addr)


# ------------------------------------------------------------------ #
# AUTH HELPERS
# ------------------------------------------------------------------ #

# ------------------------------------------------------------------ #
# PASSWORD VALIDATION
# ------------------------------------------------------------------ #

def validate_password(pw):
    """Returns (ok, message). Min 8 chars, 1 uppercase, 1 digit, 1 special."""
    if len(pw) < 8:
        return False, 'Phải ít nhất 8 ký tự'
    if not re.search(r'[A-Z]', pw):
        return False, 'Phải có ít nhất 1 chữ HOA (A-Z)'
    if not re.search(r'[0-9]', pw):
        return False, 'Phải có ít nhất 1 chữ số (0-9)'
    if not re.search(r'[^A-Za-z0-9]', pw):
        return False, 'Phải có ít nhất 1 ký tự đặc biệt (!@#$%...)'
    return True, ''


# ------------------------------------------------------------------ #
# SECURITY MIDDLEWARE
# ------------------------------------------------------------------ #

@app.before_request
def security_middleware():
    # HTTPS enforcement
    if _FORCE_HTTPS and not request.is_secure and request.method != 'OPTIONS':
        return redirect(request.url.replace('http://', 'https://', 1), code=301)

    # Session idle timeout (30 minutes)
    if session.get('user_id'):
        last_active = session.get('_last_active', 0)
        now = _dt.now().timestamp()
        if now - last_active > 1800:
            uid, uname = session.get('user_id'), session.get('username', '')
            session.clear()
            db.log_audit(uid, uname, 'SESSION_EXPIRED', 'Phiên hết hạn tự động', request.remote_addr)
            flash('Phiên đăng nhập đã hết hạn (30 phút không hoạt động).', 'error')
            return redirect(url_for('login'))
        session['_last_active'] = now
        session.modified = True

    # Force password change redirect
    if (session.get('must_change_pw') and
            request.endpoint not in ('change_password', 'logout', 'static')):
        return redirect(url_for('change_password'))


@app.after_request
def security_headers(response):
    response.headers['X-Frame-Options']        = 'DENY'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-XSS-Protection']       = '1; mode=block'
    response.headers['Referrer-Policy']        = 'strict-origin-when-cross-origin'
    if _FORCE_HTTPS:
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return response


# ------------------------------------------------------------------ #
# AUTH HELPERS
# ------------------------------------------------------------------ #

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('user_id'):
            return redirect(url_for('login', next=request.path))
        return f(*args, **kwargs)
    return decorated


def hr_required(f):
    """Only HR and admin roles can access payroll/attendance features."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('user_id'):
            return redirect(url_for('login', next=request.path))
        role = session.get('role')
        if role == 'employee':
            return redirect(url_for('portal'))   # employees have their own portal
        if role not in ('hr', 'admin'):
            return redirect(url_for('no_access'))
        return f(*args, **kwargs)
    return decorated


def delivery_required(f):
    """Delivery team page access — admin, hr, and delivery roles."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('user_id'):
            return redirect(url_for('login', next=request.path))
        role = session.get('role')
        if role == 'employee':
            return redirect(url_for('portal'))
        if role not in ('hr', 'admin', 'delivery'):
            return redirect(url_for('no_access'))
        return f(*args, **kwargs)
    return decorated


def _ensure_default_admin():
    """Create a default admin account on first run if no users exist."""
    if db.count_users() == 0:
        db.create_user(
            username='admin',
            password_hash=generate_password_hash('admin123'),
            display_name='Administrator',
            role='admin',
            must_change_password=1,
        )

_ensure_default_admin()


# ------------------------------------------------------------------ #
# TEMPLATE HELPERS
# ------------------------------------------------------------------ #

def _status_row(s):
    if 'Not Matched'  in s: return 'row-notmatch'
    if 'Missing'      in s: return 'row-missing'
    if 'Under Hours'  in s: return 'row-under'
    if 'Overtime'     in s: return 'row-ot'
    if 'After 20'     in s: return 'row-late'
    if 'Manual'       in s: return 'row-manual'
    if 'Late Clock In' in s: return 'row-latein'
    return 'row-ok'

def _status_badge(s):
    if 'Not Matched'  in s: return 'badge-notmatch'
    if 'Missing'      in s: return 'badge-missing'
    if 'Under Hours'  in s: return 'badge-under'
    if 'Overtime'     in s: return 'badge-ot'
    if 'After 20'     in s: return 'badge-late'
    if 'Manual'       in s: return 'badge-manual'
    if 'Late Clock In' in s: return 'badge-latein'
    return 'badge-ok'

app.jinja_env.globals.update(
    status_row_class=_status_row,
    status_badge_class=_status_badge,
)


# ------------------------------------------------------------------ #
# CONTEXT PROCESSORS
# ------------------------------------------------------------------ #

@app.context_processor
def inject_globals():
    review_count = 0
    if session.get('role') in ('hr', 'admin'):
        try:
            review_count = db.get_review_count()
        except Exception:
            pass
    return {
        'review_count':  review_count,
        'current_user':  session.get('display_name', ''),
        'current_role':  session.get('role', ''),
    }


# ------------------------------------------------------------------ #
# LOGIN / LOGOUT
# ------------------------------------------------------------------ #

@app.route('/login', methods=['GET', 'POST'])
def login():
    if session.get('user_id'):
        return redirect(url_for('index'))

    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user     = db.get_user_by_username(username)

        if user and check_password_hash(user['password_hash'], password):
            session.clear()
            session['user_id']        = user['id']
            session['username']       = user['username']
            session['display_name']   = user['display_name'] or user['username']
            session['role']           = user['role']
            session['department']     = user['department']
            session['employee_id']    = user.get('employee_id', '')
            session['_last_active']   = _dt.now().timestamp()
            session.permanent         = True
            if user.get('must_change_password'):
                session['must_change_pw'] = True
            db.log_audit(user['id'], user['username'], 'LOGIN', '', request.remote_addr)
            # Employee role → always go to personal portal
            if user['role'] == 'employee':
                return redirect(url_for('portal'))
            # Delivery role → delivery dashboard
            if user['role'] == 'delivery':
                return redirect(url_for('delivery'))
            raw_next = request.args.get('next', '')
            # Only allow relative redirects (no scheme/host) to prevent open redirect
            next_url = raw_next if (raw_next and raw_next.startswith('/') and not raw_next.startswith('//')) else url_for('index')
            return redirect(next_url)
        else:
            db.log_audit(0, username, 'LOGIN_FAILED', f'Sai mật khẩu', request.remote_addr)
            error = 'Sai username hoặc password.'

    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    db.log_audit(session.get('user_id', 0), session.get('username', ''), 'LOGOUT', '', request.remote_addr)
    session.clear()
    return redirect(url_for('login'))


@app.route('/no-access')
@login_required
def no_access():
    return render_template('no_access.html')


# ------------------------------------------------------------------ #
# DASHBOARD
# ------------------------------------------------------------------ #

@app.route('/')
@hr_required
def index():
    now   = _dt.now()
    stats = db.get_dashboard_stats()
    employees_leave = db.get_employee_leave_summary(now.year, now.month)
    return render_template('index.html', stats=stats,
                           employees_leave=employees_leave,
                           current_year=now.year,
                           current_month=now.month)


# ------------------------------------------------------------------ #
# UPLOAD & PROCESS
# ------------------------------------------------------------------ #

@app.route('/upload', methods=['GET', 'POST'])
@hr_required
def upload():
    if request.method == 'POST':
        f = request.files.get('file')
        if not f or f.filename == '':
            flash('Chưa chọn file.', 'error')
            return redirect(request.url)

        if not any(f.filename.lower().endswith(ext) for ext in ('.xlsx', '.xls', '.csv')):
            flash('Chỉ hỗ trợ .xlsx, .xls, .csv', 'error')
            return redirect(request.url)

        try:
            import traceback as _tb
            f.stream.seek(0)          # đảm bảo stream ở đầu file
            records = processor.parse_lark_file(f)
            if not records:
                flash('Không tìm thấy dữ liệu hợp lệ. Kiểm tra lại format file.', 'error')
                return redirect(request.url)

            employees = db.get_all_employees()
            aliases   = db.get_all_aliases()
            processed, review_count = processor.process_records(records, employees, aliases)
            db.save_attendance_batch(processed)
            db.log_audit(session['user_id'], session['username'], 'UPLOAD',
                         f'{len(processed)} records — {review_count} cần review | file: {f.filename}',
                         request.remote_addr)

            ok_count = len(processed) - review_count
            flash(
                f'✅ Xử lý xong! {len(processed)} records — '
                f'{ok_count} OK, {review_count} cần review.',
                'success',
            )
            return redirect(url_for('review') if review_count else url_for('attendance'))

        except Exception as e:
            import traceback
            flash(f'Lỗi: {e} | {traceback.format_exc()[:500]}', 'error')

    return render_template('upload.html')


# ------------------------------------------------------------------ #
# ATTENDANCE RECORDS
# ------------------------------------------------------------------ #

@app.route('/attendance/preview_export')
@hr_required
def preview_export():
    import calendar as _cal
    from datetime import date as _date, timedelta as _td

    filters = {k: v for k, v in {
        'month':  request.args.get('month'),
        'period': request.args.get('period'),
    }.items() if v}

    records   = db.get_attendance(filters)
    employees = db.get_all_employees()
    preview   = db.get_export_preview(filters)
    months    = db.get_distinct_months()

    # Build pivot for matrix view
    pivot        = {}
    emp_ids_seen = set()
    for r in records:
        pivot.setdefault(r['date'], {})[r['employee_id']] = r
        if r['employee_id']:
            emp_ids_seen.add(r['employee_id'])

    # Date range
    if filters.get('month'):
        yr, mo   = map(int, filters['month'].split('-'))
        last_day = _cal.monthrange(yr, mo)[1]
        if filters.get('period') == '1':
            d_start, d_end = _date(yr, mo, 1),  _date(yr, mo, 15)
        elif filters.get('period') == '2':
            d_start, d_end = _date(yr, mo, 16), _date(yr, mo, last_day)
        else:
            d_start, d_end = _date(yr, mo, 1),  _date(yr, mo, last_day)
        dates, d = [], d_start
        while d <= d_end:
            dates.append(d.strftime('%Y-%m-%d'))
            d += _td(days=1)
    else:
        dates = sorted(pivot.keys())

    DAYS     = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    emp_list = [e for e in employees if e['id'] in emp_ids_seen]

    from datetime import datetime as _dt
    date_meta = {}
    for d in dates:
        try:
            dt  = _dt.strptime(d, '%Y-%m-%d')
            dow = dt.weekday()
            date_meta[d] = {
                'day_name':   DAYS[dow],
                'is_weekend': dow >= 5,
                'display':    dt.strftime('%d/%m'),
                'dow':        dow,
            }
        except Exception:
            date_meta[d] = {'day_name': '', 'is_weekend': False, 'display': d, 'dow': -1}

    return render_template('preview_export.html',
                           preview=preview,
                           filters=filters,
                           months=months,
                           pivot=pivot,
                           dates=dates,
                           emp_list=emp_list,
                           date_meta=date_meta)


@app.route('/attendance')
@hr_required
def attendance():
    filters = {k: v for k, v in {
        'month':        request.args.get('month'),
        'employee_id':  request.args.get('employee_id'),
        'status':       request.args.get('status'),
        'period':       request.args.get('period'),
        'weekday_only': request.args.get('weekday_only'),
    }.items() if v}

    records   = db.get_attendance(filters)
    employees = db.get_all_employees()
    months    = db.get_distinct_months()

    return render_template('attendance.html',
                           records=records,
                           employees=employees,
                           months=months,
                           filters=filters,
                           status_list=processor.STATUS_VALUES)


@app.route('/attendance/export')
@hr_required
def export_attendance():
    filters = {k: v for k, v in {
        'month':       request.args.get('month'),
        'employee_id': request.args.get('employee_id'),
        'period':      request.args.get('period'),
    }.items() if v}

    records = db.get_attendance(filters)
    output  = io.StringIO()
    w       = csv.writer(output)
    w.writerow(['Date', 'Employee ID', 'Name', 'Department', 'Role',
                'Clock In', 'Clock Out', 'Actual Hrs', 'Break', 'Net Hrs',
                'Paid Hrs', 'Status', 'Notes', 'Action Taken', 'Approved By'])
    for r in records:
        w.writerow([r['date'], r['employee_id'], r['employee_name'],
                    r['department'], r['role'], r['clock_in'], r['clock_out'],
                    r['actual_hrs'], r['break_hrs'], r['net_hrs'], r['paid_hrs'],
                    r['status'], r['notes'], r['action_taken'], r['approved_by']])

    content = output.getvalue()
    from_d, to_d = _resolve_export_dates(filters)
    fname = _export_fname('Full', from_d, to_d, 'csv')
    _persist_export(('﻿' + content).encode('utf-8'), fname, 'Danh Sách Đầy Đủ', from_d, to_d)

    return Response(
        '﻿' + content,
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename={fname}'},
    )


@app.route('/attendance/export_summary')
@hr_required
def export_summary():
    filters = {k: v for k, v in {
        'month':       request.args.get('month'),
        'employee_id': request.args.get('employee_id'),
        'period':      request.args.get('period'),
    }.items() if v}

    records = db.get_attendance(filters)
    output  = io.StringIO()
    w       = csv.writer(output)
    w.writerow(['Ngày', 'Nhân Viên', 'Giờ Vào', 'Giờ Ra', 'Số Giờ (Net)', 'Ghi Chú'])

    for r in records:
        if r.get('action_taken'):
            by   = f" [{r['approved_by']}]" if r.get('approved_by') else ''
            note = r['action_taken'] + by
        elif r.get('status') and r['status'] != '✅ OK':
            note = r['status']
        else:
            note = ''

        w.writerow([
            r['date'],
            r['employee_name'],
            r['clock_in']  or '',
            r['clock_out'] or '',
            r['net_hrs']   if r['net_hrs'] is not None else '',
            note,
        ])

    content = output.getvalue()
    from_d, to_d = _resolve_export_dates(filters)
    fname = _export_fname('Summary', from_d, to_d, 'csv')
    _persist_export(('﻿' + content).encode('utf-8'), fname, 'Tóm Tắt Chấm Công', from_d, to_d)

    return Response(
        '﻿' + content,
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename={fname}'},
    )


@app.route('/attendance/export_matrix')
@hr_required
def export_matrix():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    filters = {k: v for k, v in {
        'month':  request.args.get('month'),
        'period': request.args.get('period'),
    }.items() if v}

    records   = db.get_attendance(filters)
    employees = db.get_all_employees()

    if not records:
        flash('Không có dữ liệu để xuất.', 'error')
        return redirect(url_for('attendance'))

    # Build pivot: {date: {emp_id: record}}
    pivot = {}
    emp_ids_seen = set()
    for r in records:
        pivot.setdefault(r['date'], {})[r['employee_id']] = r
        if r['employee_id']:
            emp_ids_seen.add(r['employee_id'])

    # Generate full date range (including weekends) when month is known
    if filters.get('month'):
        import calendar as _cal
        yr, mo = map(int, filters['month'].split('-'))
        last_day = _cal.monthrange(yr, mo)[1]
        from datetime import date as _date, timedelta as _td
        if filters.get('period') == '1':
            d_start, d_end = _date(yr, mo, 1),  _date(yr, mo, 15)
        elif filters.get('period') == '2':
            d_start, d_end = _date(yr, mo, 16), _date(yr, mo, last_day)
        else:
            d_start, d_end = _date(yr, mo, 1),  _date(yr, mo, last_day)
        dates = []
        d = d_start
        while d <= d_end:
            dates.append(d.strftime('%Y-%m-%d'))
            d += _td(days=1)
    else:
        dates = sorted(pivot.keys())
    emp_list = [e for e in employees if e['id'] in emp_ids_seen]
    n_emp    = len(emp_list)
    last_col = get_column_letter(max(9, 2 + n_emp))   # at least col I

    # ── Style helpers ──────────────────────────────────────────────
    def _fill(h): return PatternFill(start_color=h, end_color=h, fill_type='solid')
    def _font(bold=False, color='000000', size=9, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic)
    def _align(h='center', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    _thin   = Side(style='thin', color='E2E8F0')
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    def _12h(t):
        if not t: return ''
        try:
            h, m = map(int, t.split(':'))
            return f"{h%12 or 12}:{m:02d} {'AM' if h < 12 else 'PM'}"
        except Exception:
            return t

    C = {   # colors
        'hdr':  '0F172A', 'sep':  '1E293B', 'rawh': 'F8FAFC',
        'ot':   'FEF9C3', 'uh':   'DBEAFE', 'mc':   'FEE2E2',
        'late': 'FED7AA', 'off':  'F1F5F9', 'ok':   'FFFFFF',
    }
    DAYS = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']

    # ── Workbook ───────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance Matrix'

    # Row 1 – Title
    ws.merge_cells(f'A1:{last_col}1')
    c = ws['A1']
    c.value, c.font, c.alignment = (
        'Biweekly Attendance Matrix Report',
        _font(bold=True, size=14, color='0F172A'),
        _align('center'),
    )
    ws.row_dimensions[1].height = 26

    # Row 2 – Subtitle
    ws.merge_cells(f'A2:{last_col}2')
    c = ws['A2']
    c.value     = ('One-tab format: dates down the left, employee names across the top. '
                   'Each cell shows Clock In → Clock Out and net hours after 1-hour break.')
    c.font      = _font(italic=True, size=8, color='64748B')
    c.alignment = _align('center')
    ws.row_dimensions[2].height = 14

    ws.row_dimensions[3].height = 6   # spacer

    # Row 4 – Pay period + Legend
    pay_start = dates[0] if dates else ''
    ws['A4'].value = f'Pay Period Start: {pay_start}'
    ws['A4'].font  = _font(bold=True, size=9)
    ws['C4'].value, ws['C4'].font = 'Default Break', _font(size=9, color='64748B')
    ws['D4'].value, ws['D4'].font = '1.0', _font(bold=True, size=9)
    for col_l, txt, bg, fg in [
        ('F', 'OT = Review',       C['ot'],   'A16207'),
        ('G', 'UH = Under Hours',  C['uh'],   '1E40AF'),
        ('H', 'MC = Missing Clock',C['mc'],   'B91C1C'),
    ]:
        ws[f'{col_l}4'].value, ws[f'{col_l}4'].fill = txt, _fill(bg)
        ws[f'{col_l}4'].font  = _font(size=8, color=fg)
        ws[f'{col_l}4'].alignment = _align('center')
    ws.row_dimensions[4].height = 16

    ws.row_dimensions[5].height = 6   # spacer

    # Row 6 – Column headers
    for col_l, label in [('A', 'Date'), ('B', 'Day')]:
        c = ws[f'{col_l}6']
        c.value, c.fill  = label, _fill(C['hdr'])
        c.font, c.alignment, c.border = (
            _font(bold=True, color='FFFFFF', size=9), _align('center'), _border)
    for i, emp in enumerate(emp_list):
        c = ws[f'{get_column_letter(3+i)}6']
        c.value, c.fill  = emp['name'], _fill(C['hdr'])
        c.font, c.alignment, c.border = (
            _font(bold=True, color='FFFFFF', size=9), _align('center', wrap=True), _border)
    ws.row_dimensions[6].height = 30

    # ── Data rows ──────────────────────────────────────────────────
    for ri, date_str in enumerate(dates):
        rn      = 7 + ri
        dt      = _dt.strptime(date_str, '%Y-%m-%d')
        dow     = dt.weekday()
        weekend = dow >= 5

        for col_l, val, kw in [
            ('A', date_str, dict(bold=True, color='374151')),
            ('B', DAYS[dow], dict(color='6B7280')),
        ]:
            c = ws[f'{col_l}{rn}']
            c.value, c.font      = val, _font(size=9, **kw)
            c.alignment, c.border = _align('center'), _border
            if weekend:
                c.fill = _fill(C['off'])

        for i, emp in enumerate(emp_list):
            c = ws[f'{get_column_letter(3+i)}{rn}']
            c.alignment, c.border = _align('center', wrap=True), _border

            if weekend:
                c.value, c.fill = 'OFF', _fill(C['off'])
                c.font = _font(size=8, color='94A3B8')
            elif date_str in pivot and emp['id'] in pivot[date_str]:
                rec    = pivot[date_str][emp['id']]
                status = rec.get('status', '')
                ci, co = rec.get('clock_in', ''), rec.get('clock_out', '')
                net    = rec.get('net_hrs')

                if any(k in status for k in ('Missing', 'Manual', 'Not Matched')):
                    tag = 'MC' if 'Missing' in status else '—'
                    c.value, c.fill = tag, _fill(C['mc'])
                    c.font = _font(bold=True, size=9, color='DC2626')
                else:
                    if 'Overtime' in status or 'After 20' in status:
                        tag, bg, fg = '/ OT', C['ot'], '92400E'
                    elif 'Under Hours' in status:
                        tag, bg, fg = '/ UH', C['uh'], '1E40AF'
                    elif 'Late Clock In' in status:
                        tag, bg, fg = '/ Late', C['late'], 'C2410C'
                    else:
                        tag, bg, fg = '', C['ok'], '1F2937'

                    time_line = (f'{_12h(ci)} - {_12h(co)}' if ci and co
                                 else (ci or co or '—'))
                    net_line  = (f'{net}h {tag}'.strip() if net is not None
                                 else tag.strip() or '—')
                    c.value, c.fill = f'{time_line}\n{net_line}', _fill(bg)
                    c.font = _font(size=8, color=fg)
            else:
                c.fill = _fill('FAFAFA')
                c.font = _font(size=8)

        ws.row_dimensions[rn].height = 36

    # ── Raw data section ───────────────────────────────────────────
    sep_row  = 7 + len(dates) + 2
    max_raw  = max(9, 2 + n_emp)
    ws.merge_cells(f'A{sep_row}:{get_column_letter(max_raw)}{sep_row}')
    c = ws[f'A{sep_row}']
    c.value     = ('Raw Data Area — Paste Lark report here, '
                   'or let Apps Script populate this area automatically')
    c.fill      = _fill(C['sep'])
    c.font      = _font(bold=True, color='CBD5E1', size=8)
    c.alignment = _align('center')
    ws.row_dimensions[sep_row].height = 20

    hdr_row  = sep_row + 1
    raw_hdrs = ['Date', 'Employee', 'Clock In', 'Clock Out',
                'Break', 'Work Hours', 'Status', 'Display Text', 'Key']
    for j, h in enumerate(raw_hdrs):
        c = ws.cell(row=hdr_row, column=j+1, value=h)
        c.fill, c.font = _fill(C['rawh']), _font(bold=True, size=8, color='475569')
        c.alignment, c.border = _align('center'), _border
    ws.row_dimensions[hdr_row].height = 16

    STATUS_TAG = [
        ('OK',            'Normal',    '16A34A'),
        ('Overtime',      'OT Review', 'A16207'),
        ('After 20',      'OT Review', 'A16207'),
        ('Under Hours',   'UH',        'B91C1C'),
        ('Missing',       'MC',        'DC2626'),
        ('Late Clock In', 'Late',      'C2410C'),
    ]

    def _stag(s):
        for key, lbl, color in STATUS_TAG:
            if key in s:
                return lbl, color
        return '—', '94A3B8'

    data_row = hdr_row + 1
    for rec in sorted(records, key=lambda x: (x['date'], x.get('employee_name', ''))):
        ci, co = rec.get('clock_in', ''), rec.get('clock_out', '')
        net    = rec.get('net_hrs')
        slbl, sclr = _stag(rec.get('status', ''))
        display = f'{_12h(ci)} - {_12h(co)}\n{net}h' if ci and co and net is not None else ''
        key     = f"{rec['date']}|{rec.get('employee_name', '')}"
        row_vals = [rec['date'], rec.get('employee_name',''), ci, co,
                    rec.get('break_hrs', 1.0), net, slbl, display, key]
        for j, val in enumerate(row_vals):
            c = ws.cell(row=data_row, column=j+1, value=val)
            c.font      = _font(size=8, color=(sclr if j == 6 else '374151'))
            c.border    = _border
            c.alignment = _align('left', v='center', wrap=(j == 7))
        ws.row_dimensions[data_row].height = 30
        data_row += 1

    # ── Column widths ──────────────────────────────────────────────
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 5
    for i in range(n_emp):
        ws.column_dimensions[get_column_letter(3+i)].width = 20
    for j, w in enumerate([13, 18, 9, 9, 7, 10, 10, 22, 24], start=1):
        col_l = get_column_letter(j)
        cur = ws.column_dimensions[col_l].width or 0
        if w > cur:
            ws.column_dimensions[col_l].width = w

    ws.freeze_panes = ws['A7']

    # ── Output ─────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()
    from_d, to_d = _resolve_export_dates(filters)
    fname = _export_fname('Matrix', from_d, to_d, 'xlsx')
    _persist_export(excel_bytes, fname, 'Ma Trận Chấm Công', from_d, to_d)
    return Response(
        excel_bytes,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={fname}'},
    )


@app.route('/attendance/delete', methods=['POST'])
@hr_required
def delete_attendance():
    month  = request.form.get('month', '').strip()
    period = request.form.get('period', '').strip() or None

    if not month:
        flash('Cần chọn tháng để xóa.', 'error')
        return redirect(url_for('attendance'))

    deleted = db.delete_attendance(month=month, period=period)
    period_label = {'1': ' Kỳ 1 (1–15)', '2': ' Kỳ 2 (16–31)'}.get(period, '')
    flash(f'✅ Đã xóa {deleted} records của tháng {month}{period_label}.', 'success')
    return redirect(url_for('attendance'))


@app.route('/attendance/<int:att_id>/detail')
@hr_required
def attendance_detail(att_id):
    record = db.get_attendance_by_id(att_id)
    if not record:
        return jsonify({'ok': False, 'error': 'Not found'}), 404
    return jsonify({'ok': True, 'record': dict(record)})


@app.route('/attendance/delete/<int:att_id>', methods=['POST'])
@hr_required
def delete_attendance_record(att_id):
    db.delete_attendance_by_id(att_id)
    return jsonify({'ok': True})


@app.route('/attendance/sync_employees', methods=['POST'])
@hr_required
def sync_attendance_from_employees():
    updated = db.sync_all_from_employees()
    flash(f'✅ Đã sync department/role/name từ Employees cho {updated} records.', 'success')
    return redirect(url_for('attendance'))


@app.route('/attendance/delete_all', methods=['POST'])
@hr_required
def delete_all_attendance():
    deleted = db.delete_all_attendance()
    flash(f'✅ Đã xóa toàn bộ {deleted} records.', 'success')
    return redirect(url_for('attendance'))


# ------------------------------------------------------------------ #
# REVIEW QUEUE
# ------------------------------------------------------------------ #

@app.route('/review')
@hr_required
def review():
    filters = {k: v for k, v in {
        'month':       request.args.get('month'),
        'employee_id': request.args.get('employee_id'),
        'period':      request.args.get('period'),
    }.items() if v}

    records          = db.get_review_queue(filters)
    employees        = db.get_all_employees()
    months           = db.get_distinct_months()
    emp_map          = {e['id']: e for e in employees}
    unresolved_count = sum(1 for r in records if not r.get('action_taken'))

    return render_template('review.html',
                           records=records,
                           employees=employees,
                           emp_map=emp_map,
                           months=months,
                           filters=filters,
                           unresolved_count=unresolved_count)


@app.route('/review/clock_out/<int:att_id>', methods=['POST'])
@hr_required
def correct_clock_out(att_id):
    clock_out   = request.form.get('clock_out', '').strip()
    action      = request.form.get('action_taken', '').strip() or 'Clock out corrected'
    approved_by = request.form.get('approved_by', '').strip()

    if not clock_out:
        return jsonify({'ok': False, 'error': 'Cần nhập giờ clock out'}), 400

    record = db.get_attendance_by_id(att_id)
    if not record:
        return jsonify({'ok': False, 'error': 'Record không tồn tại'}), 404

    emp  = db.get_employee_by_id(record['employee_id']) if record.get('employee_id') else None
    calc = processor.recalculate_from_clockout(
        record['clock_in'], clock_out, record.get('break_hrs') or 1.0, emp
    )
    if not calc:
        return jsonify({'ok': False, 'error': 'Sai định dạng giờ (dùng HH:MM)'}), 400

    old_note = record.get('notes') or ''
    new_note = processor._note(old_note, calc['extra_note']) if calc['extra_note'] else old_note

    db.save_clock_out_correction(att_id, clock_out, calc, new_note, action, approved_by)

    return jsonify({
        'ok':        True,
        'clock_out': clock_out,
        'actual_hrs': calc['actual_hrs'],
        'net_hrs':   calc['net_hrs'],
        'paid_hrs':  calc['paid_hrs'],
        'status':    calc['status'],
    })


@app.route('/review/update/<int:att_id>', methods=['POST'])
@hr_required
def update_review(att_id):
    action      = request.form.get('action_taken', '')
    approved_by = request.form.get('approved_by', '')
    paid_hrs    = request.form.get('paid_hrs', '').strip()

    try:
        paid_val = float(paid_hrs) if paid_hrs else None
        db.update_review(att_id, action, approved_by, paid_val)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400


# ------------------------------------------------------------------ #
# EMPLOYEES & ALIASES
# ------------------------------------------------------------------ #

@app.route('/employees')
@hr_required
def employees():
    return render_template('employees.html',
                           employees=db.get_all_employees(),
                           aliases=db.get_all_aliases())


@app.route('/employees/save', methods=['POST'])
@hr_required
def save_employee():
    checked_days = request.form.getlist('work_days')   # e.g. ['0','1','2','3','4']
    work_days    = ','.join(sorted(set(checked_days))) if checked_days else '0,1,2,3,4'
    data = {
        'id':              request.form.get('id', '').strip().upper(),
        'name':            request.form.get('name', '').strip(),
        'role':            request.form.get('role', '').strip(),
        'department':      request.form.get('department', '').strip(),
        'start_time':      request.form.get('start_time', '09:00'),
        'end_time':        request.form.get('end_time', '18:00'),
        'break_hrs':       request.form.get('break_hrs', '1.0'),
        'max_hrs_day':     request.form.get('max_hrs_day', '8.0'),
        'work_days':       work_days,
        'employment_type': request.form.get('employment_type', 'full_time'),
    }
    if not data['id'] or not data['name']:
        flash('Employee ID và Name là bắt buộc.', 'error')
        return redirect(url_for('employees'))

    db.save_employee(data)
    db.sync_attendance_employee_info(data['id'], data['name'], data['department'], data['role'])
    flash(f"✅ Đã lưu nhân viên {data['name']} và cập nhật attendance records.", 'success')
    return redirect(url_for('employees'))


@app.route('/employees/delete/<emp_id>', methods=['POST'])
@hr_required
def delete_employee(emp_id):
    db.delete_employee(emp_id)
    flash('✅ Đã xóa nhân viên.', 'success')
    return redirect(url_for('employees'))


@app.route('/aliases/save', methods=['POST'])
@hr_required
def save_alias():
    lark_name   = request.form.get('lark_name', '').strip()
    employee_id = request.form.get('employee_id', '').strip()
    notes       = request.form.get('notes', '').strip()

    if not lark_name or not employee_id:
        flash('Lark Name và Employee ID là bắt buộc.', 'error')
        return redirect(url_for('employees'))

    db.save_alias(lark_name, employee_id, notes)
    flash(f'✅ Đã thêm alias "{lark_name}"', 'success')
    return redirect(url_for('employees'))


@app.route('/aliases/delete/<int:alias_id>', methods=['POST'])
@hr_required
def delete_alias(alias_id):
    db.delete_alias(alias_id)
    flash('✅ Đã xóa alias.', 'success')
    return redirect(url_for('employees'))


# ------------------------------------------------------------------ #
# TICKETS — LEAVE & LATE ARRIVAL
# ------------------------------------------------------------------ #

@app.route('/tickets')
@hr_required
def tickets():
    filters = {k: v for k, v in {
        'type':        request.args.get('type'),
        'employee_id': request.args.get('employee_id'),
        'month':       request.args.get('month'),
    }.items() if v}

    all_tickets    = db.get_all_tickets(filters)
    employees      = db.get_all_employees()
    months         = db.get_distinct_months()
    today          = _dt.now().strftime('%Y-%m-%d')
    now            = _dt.now()
    leave_summary  = db.get_employee_leave_summary(now.year, now.month)
    leave_map      = {e['id']: e for e in leave_summary}
    created_ticket = request.args.get('created', '')
    return render_template('tickets.html',
                           tickets=all_tickets,
                           employees=employees,
                           months=months,
                           filters=filters,
                           today=today,
                           leave_map=leave_map,
                           created_ticket=created_ticket,
                           current_user=session.get('display_name', ''))


@app.route('/tickets/create', methods=['POST'])
@hr_required
def create_ticket():
    ticket_type   = request.form.get('type', '').strip()
    employee_id   = request.form.get('employee_id', '').strip()
    leave_date    = request.form.get('leave_date', '').strip()
    leave_date_to = request.form.get('leave_date_to', '').strip()
    reason        = request.form.get('reason', '').strip()
    expected_time = request.form.get('expected_time', '').strip()
    approved_by   = session.get('display_name') or session.get('username', '')

    if ticket_type not in ('annual_leave', 'late_arrival'):
        flash('Loại ticket không hợp lệ.', 'error')
        return redirect(url_for('tickets'))
    if not employee_id or not leave_date:
        flash('Nhân viên và ngày là bắt buộc.', 'error')
        return redirect(url_for('tickets'))

    emp = db.get_employee_by_id(employee_id)
    if not emp:
        flash('Không tìm thấy nhân viên.', 'error')
        return redirect(url_for('tickets'))

    if emp.get('employment_type') == 'part_time':
        if ticket_type == 'annual_leave':
            flash(f'{emp["name"]} là nhân viên Part-time — không áp dụng phép năm.', 'error')
        else:
            flash(f'{emp["name"]} là nhân viên Part-time — lương tính theo giờ thực làm, không cần ticket đi trễ.', 'error')
        return redirect(url_for('tickets'))

    ticket_id = db.create_ticket({
        'type':          ticket_type,
        'employee_id':   employee_id,
        'employee_name': emp['name'],
        'leave_date':    leave_date,
        'leave_date_to': leave_date_to,
        'reason':        reason,
        'expected_time': expected_time,
        'approved_by':   approved_by,
    })

    type_label = 'Nghỉ phép năm' if ticket_type == 'annual_leave' else 'Xin đi trễ'
    date_label = f'{leave_date} → {leave_date_to}' if leave_date_to and leave_date_to != leave_date else leave_date
    flash(f'✅ Đã tạo ticket {ticket_id} — {type_label} cho {emp["name"]} ngày {date_label}.', 'success')
    return redirect(url_for('tickets', created=ticket_id))


@app.route('/tickets/delete/<int:ticket_id>', methods=['POST'])
@hr_required
def delete_ticket(ticket_id):
    db.delete_ticket(ticket_id)
    flash('✅ Đã xóa ticket.', 'success')
    return redirect(url_for('tickets'))


# ------------------------------------------------------------------ #
# MONTHLY REPORT
# ------------------------------------------------------------------ #

@app.route('/report')
@hr_required
def report():
    import calendar as _cal
    months    = db.get_distinct_months()
    from_date = request.args.get('from_date', '').strip()
    to_date   = request.args.get('to_date', '').strip()

    # Quick shortcut: month + period → auto-fill date range
    month  = request.args.get('month', '').strip()
    period = request.args.get('period', '').strip()
    if month and not from_date:
        yr, mo   = map(int, month.split('-'))
        last_day = _cal.monthrange(yr, mo)[1]
        if period == '1':
            from_date, to_date = f'{month}-01', f'{month}-15'
        elif period == '2':
            from_date, to_date = f'{month}-16', f'{month}-{last_day:02d}'
        else:
            from_date, to_date = f'{month}-01', f'{month}-{last_day:02d}'

    report_data = db.get_monthly_report(from_date or None, to_date or None)
    return render_template('report.html',
                           report=report_data,
                           months=months,
                           from_date=from_date,
                           to_date=to_date,
                           current_month=month)


@app.route('/report/employee/<emp_id>')
@hr_required
def report_employee_detail(emp_id):
    from_date = request.args.get('from_date', '') or None
    to_date   = request.args.get('to_date', '')   or None
    records   = db.get_employee_report_detail(emp_id, from_date, to_date)
    return jsonify({'ok': True, 'records': records})


@app.route('/report/export')
@hr_required
def export_report():
    from_date   = request.args.get('from_date', '') or None
    to_date     = request.args.get('to_date', '')   or None
    report_data = db.get_monthly_report(from_date, to_date)

    output = io.StringIO()
    w      = csv.writer(output)
    w.writerow(['Employee ID', 'Name', 'Department', 'Role',
                'Total Days', 'Present Days', 'Total Paid Hrs',
                'Overtime Days', 'Under Hrs Days',
                'Late In Days', 'Missing In Days', 'Missing Out Days',
                'Nghi Khong Phep', 'Nghi Phep Nam', 'Unresolved'])
    for r in report_data:
        w.writerow([
            r['employee_id'], r['employee_name'], r['department'], r['role'],
            r['total_days'], r['present_days'], r['total_paid_hrs'],
            r['overtime_days'], r['under_hrs_days'],
            r['late_in_days'], r['missing_in_days'], r['missing_out_days'],
            r['absent_unpaid_days'], r['absent_paid_days'], r['unresolved_count'],
        ])

    content = output.getvalue()
    fname = _export_fname('Report', from_date or '', to_date or '', 'csv')
    _persist_export(('﻿' + content).encode('utf-8'), fname, 'Báo Cáo Tháng', from_date or '', to_date or '')
    return Response(
        '﻿' + content,
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename={fname}'},
    )


# ------------------------------------------------------------------ #
# EXPORT HISTORY
# ------------------------------------------------------------------ #

@app.route('/exports')
@hr_required
def exports():
    records = db.get_all_exports()
    # Annotate with file-exists flag
    for r in records:
        r['exists'] = os.path.isfile(r.get('filepath', ''))
        r['size_kb'] = round(r['file_size'] / 1024, 1) if r['file_size'] else 0
    return render_template('exports.html', exports=records)


@app.route('/exports/download/<int:export_id>')
@hr_required
def download_export(export_id):
    from flask import send_file
    rec = db.get_export_by_id(export_id)
    if not rec or not os.path.isfile(rec['filepath']):
        flash('File không còn tồn tại trên hệ thống.', 'error')
        return redirect(url_for('exports'))
    ext = rec['filename'].rsplit('.', 1)[-1].lower()
    mime = ('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            if ext == 'xlsx' else 'text/csv; charset=utf-8')
    with open(rec['filepath'], 'rb') as f:
        raw = _decrypt(f.read())
    db.log_audit(session['user_id'], session['username'], 'EXPORT',
                 f'Download: {rec["filename"]}', request.remote_addr)
    return Response(raw, mimetype=mime,
                    headers={'Content-Disposition': f'attachment; filename={rec["filename"]}'})


@app.route('/exports/delete/<int:export_id>', methods=['POST'])
@hr_required
def delete_export(export_id):
    if session.get('role') != 'admin':
        flash('Chỉ Admin mới được xóa file export.', 'error')
        return redirect(url_for('exports'))
    rec = db.get_export_by_id(export_id)
    if rec:
        try:
            if os.path.isfile(rec['filepath']):
                os.remove(rec['filepath'])
        except OSError:
            pass
        db.delete_export_record(export_id)
        flash(f'✅ Đã xóa file {rec["filename"]}.', 'success')
    return redirect(url_for('exports'))


# ------------------------------------------------------------------ #
# USER MANAGEMENT (admin only)
# ------------------------------------------------------------------ #

@app.route('/users')
@hr_required
def manage_users():
    if session.get('role') != 'admin':
        return redirect(url_for('no_access'))
    return render_template('users.html', users=db.get_all_users(),
                           employees=db.get_all_employees(),
                           current_user_id=session.get('user_id'))


@app.route('/users/save', methods=['POST'])
@hr_required
def save_user():
    if session.get('role') != 'admin':
        return redirect(url_for('no_access'))

    username     = request.form.get('username', '').strip()
    display_name = request.form.get('display_name', '').strip()
    role         = request.form.get('role', 'viewer')
    department   = request.form.get('department', '').strip()
    password     = request.form.get('password', '').strip()
    employee_id  = request.form.get('employee_id', '').strip()
    email        = request.form.get('email', '').strip()

    if not username or not password:
        flash('Username và Password là bắt buộc.', 'error')
        return redirect(url_for('manage_users'))

    if role == 'employee' and not employee_id:
        flash('Role "employee" phải được gán với một nhân viên.', 'error')
        return redirect(url_for('manage_users'))

    try:
        db.create_user(username, generate_password_hash(password), display_name, role, department,
                       must_change_password=1, employee_id=employee_id)
        if email:
            user = db.get_user_by_username(username)
            if user:
                db.update_user_email(user['id'], email)
        db.log_audit(session['user_id'], session['username'], 'USER_CREATED',
                     f'Tạo user: {username} ({role})', request.remote_addr)
        flash(f'✅ Đã tạo user "{username}". Người dùng sẽ được yêu cầu đổi mật khẩu khi đăng nhập lần đầu.', 'success')
    except Exception as e:
        flash(f'Lỗi: {e}', 'error')
    return redirect(url_for('manage_users'))


@app.route('/users/update/<int:user_id>', methods=['POST'])
@hr_required
def update_user(user_id):
    if session.get('role') != 'admin':
        return redirect(url_for('no_access'))
    display_name = request.form.get('display_name', '').strip()
    department   = request.form.get('department', '').strip()
    email        = request.form.get('email', '').strip()
    new_password = request.form.get('new_password', '').strip()
    allowed_roles = ('admin', 'hr', 'employee', 'viewer', 'delivery')
    new_role     = request.form.get('role', '').strip()
    if new_role not in allowed_roles:
        new_role = None
    with db.get_db() as conn:
        if new_role:
            conn.execute(
                'UPDATE users SET display_name=?, department=?, email=?, role=? WHERE id=?',
                (display_name, department, email, new_role, user_id)
            )
        else:
            conn.execute(
                'UPDATE users SET display_name=?, department=?, email=? WHERE id=?',
                (display_name, department, email, user_id)
            )
        if new_password:
            if len(new_password) < 8:
                flash('Mật khẩu mới phải ít nhất 8 ký tự.', 'error')
                return redirect(url_for('manage_users'))
            conn.execute(
                'UPDATE users SET password_hash=? WHERE id=?',
                (generate_password_hash(new_password), user_id)
            )
    db.log_audit(session['user_id'], session['username'], 'USER_UPDATED',
                 f'Cập nhật user ID {user_id}', request.remote_addr)
    flash('✅ Đã cập nhật thông tin user.', 'success')
    return redirect(url_for('manage_users'))


@app.route('/users/delete/<int:user_id>', methods=['POST'])
@hr_required
def delete_user(user_id):
    if session.get('role') != 'admin':
        return redirect(url_for('no_access'))
    if user_id == session.get('user_id'):
        flash('Không thể tự xóa tài khoản đang dùng.', 'error')
        return redirect(url_for('manage_users'))
    target = next((u for u in db.get_all_users() if u['id'] == user_id), {})
    db.delete_user(user_id)
    db.log_audit(session['user_id'], session['username'], 'USER_DELETED',
                 f'Xóa user: {target.get("username", user_id)}', request.remote_addr)
    flash('✅ Đã xóa user.', 'success')
    return redirect(url_for('manage_users'))


@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    forced = bool(session.get('must_change_pw'))
    if request.method == 'POST':
        old_pw  = request.form.get('old_password', '')
        new_pw  = request.form.get('new_password', '').strip()
        confirm = request.form.get('confirm_password', '').strip()

        user = db.get_user_by_username(session['username'])
        if not check_password_hash(user['password_hash'], old_pw):
            flash('Mật khẩu cũ không đúng.', 'error')
            return redirect(url_for('change_password'))

        ok, msg = validate_password(new_pw)
        if not ok:
            flash(f'Mật khẩu không hợp lệ: {msg}', 'error')
            return redirect(url_for('change_password'))

        if new_pw != confirm:
            flash('Xác nhận mật khẩu không khớp.', 'error')
            return redirect(url_for('change_password'))

        db.update_user_password(session['user_id'], generate_password_hash(new_pw))
        db.set_must_change_password(session['user_id'], False)
        session.pop('must_change_pw', None)
        db.log_audit(session['user_id'], session['username'], 'PASSWORD_CHANGED', '', request.remote_addr)
        flash('✅ Đã đổi mật khẩu thành công.', 'success')
        return redirect(url_for('index') if forced else
                        (url_for('manage_users') if session['role'] == 'admin' else url_for('index')))

    return render_template('change_password.html', forced=forced)


@app.route('/users/change_password', methods=['POST'])
@login_required
def change_password_legacy():
    """Redirect legacy form submissions to new route."""
    return redirect(url_for('change_password'))


# ------------------------------------------------------------------ #
# EMPLOYEE PORTAL (employee role only)
# ------------------------------------------------------------------ #

@app.route('/portal')
@login_required
def portal():
    if session.get('role') != 'employee':
        return redirect(url_for('index'))

    emp_id = session.get('employee_id', '')
    if not emp_id:
        flash('Tài khoản này chưa được gán với nhân viên nào. Liên hệ admin.', 'error')
        return redirect(url_for('logout'))

    from_date = request.args.get('from_date', '')
    to_date   = request.args.get('to_date', '')

    # Default to current month
    today = _dt.now()
    if not from_date:
        from_date = today.replace(day=1).strftime('%Y-%m-%d')
    if not to_date:
        to_date = today.strftime('%Y-%m-%d')

    with db.get_db() as conn:
        records = conn.execute('''
            SELECT a.*, e.name AS employee_name
            FROM attendance a
            LEFT JOIN employees e ON a.employee_id = e.id
            WHERE a.employee_id = ?
              AND DATE(a.date) BETWEEN ? AND ?
            ORDER BY a.date DESC
        ''', (emp_id, from_date, to_date)).fetchall()

        emp = conn.execute('SELECT * FROM employees WHERE id = ?', (emp_id,)).fetchone()

    # Build summary stats
    total = len(records)
    ok_count      = sum(1 for r in records if '✅' in (r['status'] or ''))
    late_count     = sum(1 for r in records if 'Late' in (r['status'] or '') and 'Out' not in (r['status'] or ''))
    missing_count  = sum(1 for r in records if 'Missing' in (r['status'] or ''))
    ot_count       = sum(1 for r in records if 'Overtime' in (r['status'] or '') or 'After 20' in (r['status'] or ''))

    return render_template('portal.html',
                           records=records,
                           employee=emp,
                           from_date=from_date,
                           to_date=to_date,
                           stats={
                               'total':   total,
                               'ok':      ok_count,
                               'late':    late_count,
                               'missing': missing_count,
                               'ot':      ot_count,
                           })


@app.route('/audit')
@login_required
def audit_log():
    if session.get('role') != 'admin':
        return redirect(url_for('no_access'))
    filters = {k: v for k, v in {
        'action':   request.args.get('action'),
        'username': request.args.get('username'),
    }.items() if v}
    logs      = db.get_audit_logs(limit=1000, filters=filters)
    all_users = db.get_audit_users()
    actions   = [
        'LOGIN', 'LOGIN_FAILED', 'LOGOUT', 'SESSION_EXPIRED',
        'PASSWORD_CHANGED', 'UPLOAD', 'EXPORT', 'TICKET_CREATED',
        'REVIEW_ACTION', 'USER_CREATED', 'USER_DELETED',
        'EMPLOYEE_SAVED', 'RECORD_DELETED',
    ]
    return render_template('audit.html', logs=logs, all_users=all_users,
                           actions=actions, filters=filters)


# ------------------------------------------------------------------ #
# SYSTEM / UPDATE
# ------------------------------------------------------------------ #

_SYSTEM_CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'system_config.json')
_APP_DIR            = os.path.dirname(os.path.abspath(__file__))


# ── Email / SMTP ──────────────────────────────────────────────────────
def _send_email(to_addr, subject, html_body):
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    cfg  = _load_sys_config()
    host = cfg.get('smtp_host', '').strip()
    port = int(cfg.get('smtp_port', 587))
    user = cfg.get('smtp_user', '').strip()
    pw   = cfg.get('smtp_pass', '').strip()
    from_email = cfg.get('smtp_from_email', '').strip() or user
    from_name  = cfg.get('smtp_from_name', 'RichHR System').strip()
    use_tls    = cfg.get('smtp_tls', True)

    if not host or not user or not pw:
        raise RuntimeError('SMTP chưa được cấu hình. Vào Hệ Thống → Cài đặt Email.')

    msg = MIMEMultipart('alternative')
    msg['Subject'] = subject
    msg['From']    = f'{from_name} <{from_email}>'
    msg['To']      = to_addr
    msg.attach(MIMEText(html_body, 'html', 'utf-8'))

    with smtplib.SMTP(host, port, timeout=15) as server:
        if use_tls:
            server.starttls()
        server.login(user, pw)
        server.sendmail(from_email, [to_addr], msg.as_string())

def _load_sys_config():
    import json
    try:
        with open(_SYSTEM_CONFIG_PATH, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {'github_repo': '', 'branch': 'main', 'version': '1.0.0', 'last_update': ''}

def _save_sys_config(cfg):
    import json
    with open(_SYSTEM_CONFIG_PATH, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)


# ── Forgot / Reset password ──────────────────────────────────────────
@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'GET':
        return render_template('forgot_password.html')

    email = request.form.get('email', '').strip()
    if not email:
        return render_template('forgot_password.html', error='Vui lòng nhập email.')

    user = db.get_user_by_email(email)
    # Always show success — don't reveal whether email exists
    if user:
        import secrets as _sec
        token      = _sec.token_urlsafe(48)
        expires_at = (_dt.now() + timedelta(hours=1)).strftime('%Y-%m-%d %H:%M:%S')
        db.create_reset_token(token, user['username'], expires_at)
        reset_url  = url_for('reset_password', token=token, _external=True)
        html_body  = f'''
<div style="font-family:Inter,sans-serif;max-width:520px;margin:0 auto;padding:32px 24px">
  <div style="background:#2563eb;border-radius:14px;padding:28px 32px;text-align:center;margin-bottom:24px">
    <div style="font-size:2rem;margin-bottom:8px">🔐</div>
    <h1 style="color:#fff;font-size:1.3rem;margin:0;font-weight:800">Đặt Lại Mật Khẩu</h1>
    <p style="color:#bfdbfe;font-size:.85rem;margin:6px 0 0">RichHR Attendance System</p>
  </div>
  <p style="color:#334155;font-size:.95rem">Xin chào <strong>{user["display_name"] or user["username"]}</strong>,</p>
  <p style="color:#475569;font-size:.9rem;line-height:1.6">
    Chúng tôi nhận được yêu cầu đặt lại mật khẩu cho tài khoản <strong>{user["username"]}</strong>.
    Nhấn nút bên dưới để tiếp tục (liên kết hết hạn sau <strong>1 giờ</strong>).
  </p>
  <div style="text-align:center;margin:28px 0">
    <a href="{reset_url}"
       style="background:#2563eb;color:#fff;text-decoration:none;padding:13px 32px;
              border-radius:10px;font-weight:700;font-size:.95rem;display:inline-block">
      Đặt Lại Mật Khẩu
    </a>
  </div>
  <p style="color:#94a3b8;font-size:.78rem;line-height:1.6">
    Nếu bạn không yêu cầu, hãy bỏ qua email này. Mật khẩu sẽ không thay đổi.<br>
    Liên kết: <a href="{reset_url}" style="color:#2563eb">{reset_url}</a>
  </p>
  <hr style="border:none;border-top:1px solid #e2e8f0;margin:20px 0">
  <p style="color:#cbd5e1;font-size:.75rem;text-align:center">Rich Payment Solutions · RichHR System</p>
</div>'''
        try:
            _send_email(email, 'Đặt lại mật khẩu RichHR', html_body)
        except Exception as e:
            return render_template('forgot_password.html',
                                   error=f'Không gửi được email: {e}. Liên hệ admin.')

    return render_template('forgot_password.html', sent=True)


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    rec = db.get_reset_token(token)
    if not rec:
        return render_template('reset_password.html', invalid=True)

    # Check expiry
    from datetime import datetime as _dtt
    if _dtt.strptime(rec['expires_at'], '%Y-%m-%d %H:%M:%S') < _dtt.now():
        return render_template('reset_password.html', expired=True)

    if request.method == 'GET':
        return render_template('reset_password.html', token=token, username=rec['username'])

    pw1 = request.form.get('password', '').strip()
    pw2 = request.form.get('password2', '').strip()
    if len(pw1) < 8:
        return render_template('reset_password.html', token=token, username=rec['username'],
                               error='Mật khẩu phải ít nhất 8 ký tự.')
    if pw1 != pw2:
        return render_template('reset_password.html', token=token, username=rec['username'],
                               error='Hai mật khẩu không khớp.')

    user = db.get_user_by_username(rec['username'])
    if not user:
        return render_template('reset_password.html', invalid=True)

    db.update_user_password(user['id'], generate_password_hash(pw1))
    db.consume_reset_token(token)
    return render_template('reset_password.html', success=True)


@app.route('/system')
@login_required
def system():
    if session.get('role') != 'admin':
        flash('Chỉ admin mới xem được trang này.', 'error')
        return redirect(url_for('index'))
    import sys, platform, sqlite3
    cfg = _load_sys_config()
    db_path = os.path.join(_APP_DIR, 'attendance.db')
    db_size = round(os.path.getsize(db_path) / 1024, 1) if os.path.exists(db_path) else 0
    return render_template('system.html',
                           cfg=cfg,
                           py_version=sys.version.split()[0],
                           platform=platform.system() + ' ' + platform.release(),
                           db_size=db_size,
                           current_user=session.get('display_name', ''))


@app.route('/system/config', methods=['POST'])
@login_required
def system_config_save():
    if session.get('role') != 'admin':
        return jsonify({'ok': False, 'error': 'Forbidden'}), 403
    cfg = _load_sys_config()
    cfg['github_repo'] = request.form.get('github_repo', '').strip()
    cfg['branch']      = request.form.get('branch', 'main').strip() or 'main'
    _save_sys_config(cfg)
    flash('✅ Đã lưu cấu hình GitHub.', 'success')
    return redirect(url_for('system'))


@app.route('/system/smtp', methods=['POST'])
@login_required
def system_smtp_save():
    if session.get('role') != 'admin':
        return jsonify({'ok': False, 'error': 'Forbidden'}), 403
    cfg = _load_sys_config()
    cfg['smtp_host']       = request.form.get('smtp_host', '').strip()
    cfg['smtp_port']       = int(request.form.get('smtp_port', 587) or 587)
    cfg['smtp_user']       = request.form.get('smtp_user', '').strip()
    cfg['smtp_from_email'] = request.form.get('smtp_from_email', '').strip()
    cfg['smtp_from_name']  = request.form.get('smtp_from_name', 'RichHR System').strip()
    cfg['smtp_tls']        = request.form.get('smtp_tls') == '1'
    # Only overwrite password if provided
    new_pass = request.form.get('smtp_pass', '').strip()
    if new_pass:
        cfg['smtp_pass'] = new_pass
    _save_sys_config(cfg)
    flash('✅ Đã lưu cấu hình SMTP.', 'success')
    return redirect(url_for('system'))


@app.route('/system/smtp-test', methods=['POST'])
@login_required
def system_smtp_test():
    if session.get('role') != 'admin':
        return jsonify({'ok': False, 'error': 'Forbidden'}), 403
    to_addr = request.form.get('test_email', '').strip()
    if not to_addr:
        return jsonify({'ok': False, 'error': 'Nhập email nhận thử nghiệm.'})
    try:
        _send_email(to_addr, '✅ Test email — RichHR System',
                    '<p style="font-family:sans-serif">SMTP hoạt động tốt! Hệ thống RichHR đã kết nối email thành công.</p>')
        return jsonify({'ok': True, 'message': f'Đã gửi test email tới {to_addr}'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


_GIT_EXE = r'C:\Program Files\Git\bin\git.exe'

def _find_git():
    import shutil
    return shutil.which('git') or (_GIT_EXE if os.path.exists(_GIT_EXE) else None)


@app.route('/system/update', methods=['POST'])
@login_required
def system_update():
    if session.get('role') != 'admin':
        return jsonify({'ok': False, 'error': 'Forbidden'}), 403

    import subprocess, threading, time, sys

    cfg    = _load_sys_config()
    branch = cfg.get('branch', 'main').strip() or 'main'
    steps  = []
    git    = _find_git()

    if not git:
        return jsonify({'ok': False, 'error': 'Git chưa được cài. Liên hệ admin.'})

    # ── git fetch + pull ──────────────────────────────────────────
    try:
        steps.append('Đang kết nối GitHub…')

        fetch = subprocess.run(
            [git, 'fetch', 'origin', branch],
            capture_output=True, text=True, cwd=_APP_DIR, timeout=30
        )
        if fetch.returncode != 0:
            return jsonify({'ok': False, 'error': fetch.stderr.strip() or 'git fetch thất bại', 'steps': steps})
        steps.append(f'Fetch OK từ origin/{branch}')

        # Count incoming commits
        log = subprocess.run(
            [git, 'log', f'HEAD..origin/{branch}', '--oneline'],
            capture_output=True, text=True, cwd=_APP_DIR
        )
        commits = [l for l in log.stdout.strip().splitlines() if l]
        if not commits:
            return jsonify({'ok': True, 'steps': steps,
                            'message': 'Hệ thống đang chạy phiên bản mới nhất — không có gì để cập nhật.',
                            'uptodate': True})

        steps.append(f'{len(commits)} commit mới: ' + ' · '.join(c[:60] for c in commits[:3]))

        pull = subprocess.run(
            [git, '-c', 'core.autocrlf=false', 'pull', 'origin', branch, '--ff-only'],
            capture_output=True, text=True, cwd=_APP_DIR, timeout=60
        )
        if pull.returncode != 0:
            return jsonify({'ok': False, 'error': pull.stderr.strip() or 'git pull thất bại', 'steps': steps})

        steps.append('Pull thành công — ' + (pull.stdout.strip().splitlines()[-1] if pull.stdout.strip() else 'OK'))

    except subprocess.TimeoutExpired:
        return jsonify({'ok': False, 'error': 'Timeout khi kết nối GitHub. Kiểm tra internet.', 'steps': steps})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'steps': steps})

    # ── Save timestamp & restart ──────────────────────────────────
    cfg['last_update'] = _dt.now().strftime('%Y-%m-%d %H:%M:%S')
    _save_sys_config(cfg)
    steps.append('Đang khởi động lại server…')

    def _restart():
        time.sleep(1.5)
        os.execv(sys.executable, [sys.executable] + sys.argv)

    threading.Thread(target=_restart, daemon=True).start()

    return jsonify({'ok': True, 'steps': steps,
                    'message': 'Cập nhật thành công! Server đang khởi động lại…'})


# ------------------------------------------------------------------ #
# DELIVERY DEPARTMENT — INVENTORY, DELIVERIES, EMAIL LABELS, RETURNS
# ------------------------------------------------------------------ #

_DELIVERY_STATUS_LABEL = {
    'scheduled': 'Đã lên lịch',
    'en_route':  'Đang đi giao',
    'delivered': 'Đã giao & setup',
    'returned':  'Đã thu hồi máy',
}


def _delivery_label_html(d):
    """Build the HTML 'shipping label' emailed to the customer."""
    cond_txt = {'new': 'Máy mới', 'used': 'Máy đã qua sử dụng'}.get(d.get('machine_condition'), '')
    sched    = d.get('scheduled_date') or 'Sẽ thông báo'
    return f'''
<div style="font-family:Inter,Arial,sans-serif;max-width:560px;margin:0 auto;padding:32px 24px">
  <div style="background:#0f172a;border-radius:14px;padding:26px 30px;margin-bottom:22px">
    <div style="color:#93c5fd;font-size:.72rem;letter-spacing:1px;text-transform:uppercase">Rich Payment Solutions</div>
    <h1 style="color:#fff;font-size:1.35rem;margin:6px 0 0;font-weight:800">Phiếu Giao Máy</h1>
    <div style="color:#cbd5e1;font-size:.9rem;margin-top:6px">Mã đơn: <strong style="color:#fff">{d['delivery_code']}</strong></div>
  </div>
  <p style="color:#334155;font-size:.95rem">Xin chào <strong>{d['customer_name']}</strong>,</p>
  <p style="color:#475569;font-size:.9rem;line-height:1.6">
    Đơn giao máy của quý khách đã được tạo. Vui lòng kiểm tra thông tin bên dưới khi nhận máy.
  </p>
  <table style="width:100%;border-collapse:collapse;margin:20px 0;font-size:.88rem">
    <tr><td style="padding:9px 0;color:#94a3b8;width:42%">Khách hàng</td><td style="padding:9px 0;color:#0f172a;font-weight:600">{d['customer_name']}</td></tr>
    <tr><td style="padding:9px 0;color:#94a3b8">Địa chỉ giao</td><td style="padding:9px 0;color:#0f172a">{d.get('customer_address') or '—'}</td></tr>
    <tr><td style="padding:9px 0;color:#94a3b8">Điện thoại</td><td style="padding:9px 0;color:#0f172a">{d.get('customer_phone') or '—'}</td></tr>
    <tr><td style="padding:9px 0;color:#94a3b8">Thiết bị</td><td style="padding:9px 0;color:#0f172a;font-weight:600">{d.get('machine_model') or '—'}</td></tr>
    <tr><td style="padding:9px 0;color:#94a3b8">Serial máy</td><td style="padding:9px 0;color:#0f172a"><code>{d.get('machine_serial') or '—'}</code> <span style="color:#64748b">{cond_txt}</span></td></tr>
    <tr><td style="padding:9px 0;color:#94a3b8">Ngày giao dự kiến</td><td style="padding:9px 0;color:#0f172a;font-weight:600">{sched}</td></tr>
    <tr><td style="padding:9px 0;color:#94a3b8">Nhân viên giao</td><td style="padding:9px 0;color:#0f172a">{d.get('assigned_to') or '—'}</td></tr>
  </table>
  <div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:10px;padding:14px 16px;margin:18px 0">
    <div style="color:#1e40af;font-size:.82rem;line-height:1.6">
      Nhân viên kỹ thuật sẽ liên hệ và đến lắp đặt, cài đặt máy tận nơi. Vui lòng giữ lại phiếu này để đối chiếu khi cần thu hồi/đổi trả máy.
    </div>
  </div>
  <hr style="border:none;border-top:1px solid #e2e8f0;margin:20px 0">
  <p style="color:#cbd5e1;font-size:.75rem;text-align:center">Rich Payment Solutions · Delivery Team</p>
</div>'''


@app.route('/delivery')
@delivery_required
def delivery():
    filters = {k: v for k, v in {
        'status': request.args.get('status'),
        'date':   request.args.get('date'),
        'q':      request.args.get('q'),
    }.items() if v}
    deliveries = db.get_all_deliveries(filters)
    stats      = db.get_delivery_stats()
    in_stock   = db.get_inventory_in_stock()
    today      = _dt.now().strftime('%Y-%m-%d')
    now_hm     = _dt.now().strftime('%H:%M')
    return render_template('delivery.html',
                           deliveries=deliveries,
                           stats=stats,
                           in_stock=in_stock,
                           filters=filters,
                           today=today,
                           now_hm=now_hm,
                           status_labels=_DELIVERY_STATUS_LABEL,
                           current_user=session.get('display_name', ''))


@app.route('/delivery/create', methods=['POST'])
@delivery_required
def delivery_create():
    customer_name = request.form.get('customer_name', '').strip()
    if not customer_name:
        flash('Tên khách hàng là bắt buộc.', 'error')
        return redirect(url_for('delivery'))

    serial = request.form.get('machine_serial', '').strip()
    if serial:
        machine = db.get_inventory_by_serial(serial)
        if not machine:
            flash('Máy không tồn tại trong kho.', 'error')
            return redirect(url_for('delivery'))
        if machine['status'] != 'in_stock':
            flash(f'Máy {serial} không còn trong kho (trạng thái: {machine["status"]}).', 'error')
            return redirect(url_for('delivery'))

    code = db.create_delivery({
        'customer_name':    customer_name,
        'customer_email':   request.form.get('customer_email', '').strip(),
        'customer_phone':   request.form.get('customer_phone', '').strip(),
        'customer_address': request.form.get('customer_address', '').strip(),
        'machine_serial':   serial,
        'scheduled_date':   request.form.get('scheduled_date', '').strip(),
        'assigned_to':      request.form.get('assigned_to', '').strip() or session.get('display_name', ''),
        'notes':            request.form.get('notes', '').strip(),
    })
    db.log_audit(session['user_id'], session['username'], 'DELIVERY_CREATED',
                 f'{code} → {customer_name}', request.remote_addr)
    flash(f'✅ Đã tạo đơn giao {code} cho {customer_name}.', 'success')
    return redirect(url_for('delivery'))


@app.route('/delivery/<int:delivery_id>/log', methods=['POST'])
@delivery_required
def delivery_log(delivery_id):
    """Log departure/arrival/setup times, location, and status."""
    rec = db.get_delivery_by_id(delivery_id)
    if not rec:
        return jsonify({'ok': False, 'error': 'Đơn không tồn tại'}), 404
    fields = {}
    for k in ('depart_time', 'arrive_time', 'setup_time', 'location', 'status', 'assigned_to'):
        v = request.form.get(k)
        if v is not None and v.strip() != '':
            fields[k] = v.strip()
    db.update_delivery_log(delivery_id, fields)
    db.log_audit(session['user_id'], session['username'], 'DELIVERY_LOG',
                 f'{rec["delivery_code"]} | ' + ', '.join(f'{k}={v}' for k, v in fields.items()),
                 request.remote_addr)
    if request.form.get('ajax'):
        return jsonify({'ok': True})
    flash(f'✅ Đã cập nhật nhật ký giao hàng cho {rec["delivery_code"]}.', 'success')
    return redirect(url_for('delivery'))


@app.route('/delivery/<int:delivery_id>/send-label', methods=['POST'])
@delivery_required
def delivery_send_label(delivery_id):
    rec = db.get_delivery_by_id(delivery_id)
    if not rec:
        return jsonify({'ok': False, 'error': 'Đơn không tồn tại'}), 404
    if not rec.get('customer_email'):
        return jsonify({'ok': False, 'error': 'Đơn này chưa có email khách hàng.'})
    try:
        _send_email(rec['customer_email'],
                    f'Phiếu giao máy {rec["delivery_code"]} — Rich Payment Solutions',
                    _delivery_label_html(rec))
        db.mark_label_sent(delivery_id)
        db.log_audit(session['user_id'], session['username'], 'DELIVERY_LABEL_SENT',
                     f'{rec["delivery_code"]} → {rec["customer_email"]}', request.remote_addr)
        return jsonify({'ok': True, 'message': f'Đã gửi phiếu giao tới {rec["customer_email"]}'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/delivery/<int:delivery_id>/return', methods=['POST'])
@delivery_required
def delivery_return(delivery_id):
    rec = db.get_delivery_by_id(delivery_id)
    if not rec:
        flash('Đơn không tồn tại.', 'error')
        return redirect(url_for('delivery'))
    reason = request.form.get('return_reason', '').strip()
    db.mark_delivery_returned(delivery_id, reason)
    db.log_audit(session['user_id'], session['username'], 'DELIVERY_RETURNED',
                 f'{rec["delivery_code"]} | {reason}', request.remote_addr)
    flash(f'✅ Đã ghi nhận thu hồi máy của đơn {rec["delivery_code"]}.', 'success')
    return redirect(url_for('delivery'))


@app.route('/delivery/<int:delivery_id>/delete', methods=['POST'])
@delivery_required
def delivery_delete(delivery_id):
    rec = db.get_delivery_by_id(delivery_id)
    db.delete_delivery(delivery_id)
    if rec:
        db.log_audit(session['user_id'], session['username'], 'DELIVERY_DELETED',
                     rec['delivery_code'], request.remote_addr)
    flash('✅ Đã xóa đơn giao.', 'success')
    return redirect(url_for('delivery'))


# ── Inventory ─────────────────────────────────────────────────────────

@app.route('/delivery/inventory')
@delivery_required
def delivery_inventory():
    filters = {k: v for k, v in {
        'condition': request.args.get('condition'),
        'status':    request.args.get('status'),
        'q':         request.args.get('q'),
    }.items() if v}
    items = db.get_all_inventory(filters)
    stats = db.get_inventory_stats()
    return render_template('delivery_inventory.html',
                           items=items, stats=stats, filters=filters)


@app.route('/delivery/inventory/save', methods=['POST'])
@delivery_required
def delivery_inventory_save():
    serial = request.form.get('serial_no', '').strip()
    if not serial:
        flash('Serial máy là bắt buộc.', 'error')
        return redirect(url_for('delivery_inventory'))
    db.save_inventory({
        'serial_no': serial,
        'model':     request.form.get('model', '').strip(),
        'condition': request.form.get('condition', 'new'),
        'status':    request.form.get('status', 'in_stock'),
        'notes':     request.form.get('notes', '').strip(),
    })
    db.log_audit(session['user_id'], session['username'], 'INVENTORY_SAVED',
                 f'{serial}', request.remote_addr)
    flash(f'✅ Đã lưu máy {serial}.', 'success')
    return redirect(url_for('delivery_inventory'))


@app.route('/delivery/inventory/delete/<int:inv_id>', methods=['POST'])
@delivery_required
def delivery_inventory_delete(inv_id):
    db.delete_inventory(inv_id)
    flash('✅ Đã xóa máy khỏi kho.', 'success')
    return redirect(url_for('delivery_inventory'))


# ------------------------------------------------------------------ #
# LARK API SYNC
# ------------------------------------------------------------------ #

def _lark_get_token(base_url, app_id, app_secret):
    url     = f'{base_url}/open-apis/auth/v3/tenant_access_token/internal'
    payload = _json.dumps({'app_id': app_id, 'app_secret': app_secret}).encode()
    req     = urllib.request.Request(url, data=payload,
                                     headers={'Content-Type': 'application/json'})
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = _json.loads(resp.read())
            if data.get('code') == 0:
                return data.get('tenant_access_token', '')
    except Exception:
        return None
    return None


def _lark_fetch_records(base_url, token, user_ids, date_from_int, date_to_int):
    url     = f'{base_url}/open-apis/attendance/v1/user_tasks/query'
    payload = _json.dumps({
        'user_ids':         user_ids,
        'check_date_from':  date_from_int,
        'check_date_to':    date_to_int,
        'need_overtime_result': False,
    }).encode()
    headers = {'Content-Type': 'application/json', 'Authorization': f'Bearer {token}'}
    req = urllib.request.Request(url, data=payload, headers=headers)
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = _json.loads(resp.read())
        if data.get('code') != 0:
            raise RuntimeError(f'Lark API lỗi {data.get("code")}: {data.get("msg")}')
        return data.get('data', {}).get('user_task_results', [])


def _calc_att_metrics(clock_in, clock_out, emp):
    start_t  = (emp.get('start_time') or '09:00')[:5]
    break_h  = float(emp.get('break_hrs', 1.0))
    max_h    = float(emp.get('max_hrs_day', 8.0))

    if not clock_in and not clock_out:
        return {'actual_hrs': 0, 'net_hrs': 0, 'paid_hrs': 0,
                'status': '⚠️ Missing Clock In / Clock Out'}
    if not clock_in:
        return {'actual_hrs': 0, 'net_hrs': 0, 'paid_hrs': 0, 'status': '⚠️ Missing Clock In'}
    if not clock_out:
        return {'actual_hrs': 0, 'net_hrs': 0, 'paid_hrs': 0, 'status': '⚠️ Missing Clock Out'}
    try:
        ci = _dt.strptime(clock_in[:5], '%H:%M')
        co = _dt.strptime(clock_out[:5], '%H:%M')
        st = _dt.strptime(start_t, '%H:%M')
        secs = (co - ci).total_seconds()
        if secs < 0:
            secs += 86400
        actual_hrs = secs / 3600
        net_hrs    = max(0.0, actual_hrs - break_h)
        paid_hrs   = round(min(net_hrs, max_h), 2)
        late_min   = (ci - st).total_seconds() / 60
        if late_min > 15:
            status = '⚠️ Late Clock In'
        elif net_hrs < max_h * 0.75:
            status = '⚠️ Under Hours'
        elif actual_hrs > max_h + break_h + 0.5:
            status = '⚠️ Overtime'
        else:
            status = '✅ OK'
        return {'actual_hrs': round(actual_hrs, 2), 'net_hrs': round(net_hrs, 2),
                'paid_hrs': paid_hrs, 'status': status}
    except Exception:
        return {'actual_hrs': 0, 'net_hrs': 0, 'paid_hrs': 0, 'status': '⚠️ Error'}


def _lark_sync_run(date_from_str, date_to_str, triggered_by='manual'):
    from datetime import date as _date
    cfg        = _load_sys_config()
    app_id     = cfg.get('lark_app_id', '').strip()
    app_secret = cfg.get('lark_app_secret', '').strip()
    base_url   = cfg.get('lark_base_url', 'https://open.feishu.cn').rstrip('/')
    logs       = []

    if not app_id or not app_secret:
        return {'ok': False, 'error': 'Chưa cấu hình Lark App ID / App Secret', 'logs': []}

    logs.append('→ Đang lấy access token từ Lark...')
    token = _lark_get_token(base_url, app_id, app_secret)
    if not token:
        db.add_lark_sync_log(triggered_by, date_from_str, date_to_str, 0,
                             'Lỗi: không lấy được token', 'error')
        return {'ok': False, 'error': 'Không lấy được token — kiểm tra App ID / Secret', 'logs': logs}
    logs.append('✓ Token OK')

    employees = db.get_all_employees()
    mapped    = [e for e in employees if e.get('lark_user_id', '').strip()]
    if not mapped:
        return {'ok': False,
                'error': 'Chưa có nhân viên nào được mapping Lark User ID', 'logs': logs}

    user_map  = {e['lark_user_id']: e for e in mapped}
    lark_ids  = list(user_map.keys())
    d_from_i  = int(_date.fromisoformat(date_from_str).strftime('%Y%m%d'))
    d_to_i    = int(_date.fromisoformat(date_to_str).strftime('%Y%m%d'))

    logs.append(f'→ Fetch records cho {len(lark_ids)} NV ({date_from_str} → {date_to_str})...')
    try:
        tasks = _lark_fetch_records(base_url, token, lark_ids, d_from_i, d_to_i)
    except Exception as e:
        db.add_lark_sync_log(triggered_by, date_from_str, date_to_str, 0, str(e), 'error')
        return {'ok': False, 'error': str(e), 'logs': logs}

    logs.append(f'✓ Nhận được {len(tasks)} records từ Lark')
    batch = []

    for task in tasks:
        uid = task.get('user_id', '')
        emp = user_map.get(uid)
        if not emp:
            continue
        day = str(task.get('day', ''))
        if len(day) != 8:
            continue
        date_str     = f'{day[:4]}-{day[4:6]}-{day[6:]}'
        records      = task.get('records', [])
        ci_ts = co_ts = None
        for r in records:
            sub = r.get('user_sub_type', 0)
            ts  = r.get('check_time', '')
            if not ts:
                continue
            ts_i = int(ts)
            if sub == 1 and ci_ts is None:
                ci_ts = ts_i
            elif sub == 2:
                co_ts = ts_i
        ci_str = _dt.fromtimestamp(ci_ts).strftime('%H:%M') if ci_ts else ''
        co_str = _dt.fromtimestamp(co_ts).strftime('%H:%M') if co_ts else ''
        calc   = _calc_att_metrics(ci_str, co_str, emp)
        batch.append({
            'date':          date_str,
            'employee_id':   emp['id'],
            'employee_name': emp['name'],
            'lark_name':     emp['name'],
            'department':    emp.get('department', ''),
            'role':          emp.get('role', ''),
            'clock_in':      ci_str,
            'clock_out':     co_str,
            'actual_hrs':    calc['actual_hrs'],
            'break_hrs':     emp.get('break_hrs', 1.0),
            'net_hrs':       calc['net_hrs'],
            'paid_hrs':      calc['paid_hrs'],
            'status':        calc['status'],
            'notes':         'Lark API Sync',
        })
        logs.append(f'  ✓ {emp["name"]} [{date_str}] {ci_str or "—"}→{co_str or "—"} {calc["status"]}')

    if batch:
        db.save_attendance_batch(batch)
    db.add_lark_sync_log(triggered_by, date_from_str, date_to_str,
                         len(batch), '\n'.join(logs), 'ok')
    return {'ok': True, 'processed': len(batch), 'logs': logs}


# ── Background auto-sync loop ─────────────────────────────────────────

def _lark_auto_sync_loop():
    import time
    while True:
        try:
            cfg = _load_sys_config()
            if cfg.get('lark_auto_sync'):
                from datetime import date as _date
                today = _date.today().isoformat()
                _lark_sync_run(today, today, triggered_by='auto')
                interval = int(cfg.get('lark_sync_interval', 15)) * 60
            else:
                interval = 60
        except Exception:
            interval = 60
        time.sleep(interval)

_lark_bg = _threading.Thread(target=_lark_auto_sync_loop, daemon=True)
_lark_bg.start()


# ── Lark sync routes ──────────────────────────────────────────────────

@app.route('/lark-sync')
@login_required
@hr_required
def lark_sync_page():
    cfg   = _load_sys_config()
    emps  = db.get_all_employees()
    slogs = db.get_lark_sync_logs(30)
    from datetime import date as _date
    today = _date.today().isoformat()
    return render_template('lark_sync.html', employees=emps,
                           lark_cfg=cfg, sync_logs=slogs, today=today)


@app.route('/lark/config', methods=['POST'])
@login_required
@hr_required
def lark_config_save():
    cfg = _load_sys_config()
    cfg['lark_app_id']        = request.form.get('lark_app_id', '').strip()
    cfg['lark_app_secret']    = request.form.get('lark_app_secret', '').strip()
    cfg['lark_base_url']      = request.form.get('lark_base_url', 'https://open.feishu.cn').strip()
    cfg['lark_auto_sync']     = request.form.get('lark_auto_sync') == 'on'
    cfg['lark_sync_interval'] = int(request.form.get('lark_sync_interval', 15) or 15)
    _save_sys_config(cfg)
    flash('Đã lưu cấu hình Lark API', 'success')
    return redirect(url_for('lark_sync_page'))


@app.route('/lark/mapping', methods=['POST'])
@login_required
@hr_required
def lark_save_mapping():
    for key, val in request.form.items():
        if key.startswith('lark_uid_'):
            emp_id = key[len('lark_uid_'):]
            db.save_lark_user_id(emp_id, val.strip())
    flash('Đã lưu mapping Lark User ID', 'success')
    return redirect(url_for('lark_sync_page'))


@app.route('/lark/test-token', methods=['POST'])
@login_required
@hr_required
def lark_test_token():
    cfg        = _load_sys_config()
    app_id     = cfg.get('lark_app_id', '').strip()
    app_secret = cfg.get('lark_app_secret', '').strip()
    base_url   = cfg.get('lark_base_url', 'https://open.feishu.cn').rstrip('/')
    if not app_id or not app_secret:
        return jsonify({'ok': False, 'error': 'Chưa lưu App ID / Secret'})
    token = _lark_get_token(base_url, app_id, app_secret)
    if token:
        return jsonify({'ok': True, 'message': 'Kết nối Lark thành công!'})
    return jsonify({'ok': False, 'error': 'Lỗi kết nối — kiểm tra App ID / Secret và quyền app'})


@app.route('/lark/sync', methods=['POST'])
@login_required
@hr_required
def lark_sync_manual():
    from datetime import date as _date
    date_from = request.form.get('date_from', _date.today().isoformat())
    date_to   = request.form.get('date_to',   _date.today().isoformat())
    result    = _lark_sync_run(date_from, date_to,
                               triggered_by=session.get('username', 'manual'))
    return jsonify(result)


# ------------------------------------------------------------------ #

if __name__ == '__main__':
    import sys, io as _io
    sys.stdout = _io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = _io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    print('\n' + '='*50)
    print('  Attendance Management System')
    print('  Mo trinh duyet: http://localhost:5000')
    print('  Nhan Ctrl+C de dung')
    print('='*50 + '\n')
    app.run(debug=False, host='127.0.0.1', port=5000)
